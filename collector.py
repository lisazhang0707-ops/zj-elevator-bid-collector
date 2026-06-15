"""
浙江省电梯招投标信息自动采集脚本 v1.0
数据源：浙江省公共资源交易平台 API (ggzy.zj.gov.cn)
输出：Excel文件（带颜色分类）

使用方法：
  pip install -r requirements.txt
  python collector.py

  # 可选参数
  python collector.py --days 3          # 只采集最近3天的
  python collector.py --output ./output # 指定输出目录
  python collector.py --keywords "电梯,自动扶梯"  # 自定义关键词
"""

import requests
import json
import re
import os
import argparse
from datetime import datetime, timedelta
from pathlib import Path

# ============ 默认配置 ============
DEFAULT_OUTPUT_DIR = Path("./output")
API_URL = "https://ggzy.zj.gov.cn/inteligentsearch/rest/esinteligentsearch/getFullTextDataNew"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Content-Type": "application/json;charset=UTF-8",
    "Referer": "https://ggzy.zj.gov.cn/jyxxgk/list.html",
    "Origin": "https://ggzy.zj.gov.cn",
}

# 搜索关键词组
DEFAULT_KEYWORDS = ["电梯", "电梯采购", "电梯维保", "电梯维修", "电梯改造", "加装电梯"]

# 业务类型分类规则
BUSINESS_RULES = {
    "新梯采购": ["电梯设备采购", "电梯采购", "电梯设备供货", "电梯设备安装", "电梯购置", "电梯供货"],
    "维保服务": ["维保", "维护保养", "日常维护", "年度维保", "保养服务"],
    "维修改造": ["电梯维修", "电梯改造", "电梯大修", "更新改造", "更换"],
    "旧梯加装": ["加装电梯", "既有建筑", "旧楼加装", "老旧小区.*电梯"],
}

# 类别代码映射（从API返回的infoa字段）
CATEGORY_MAP = {
    "A01": "招标公告",
    "A02": "资格预审公告",
    "A03": "项目登记信息",
    "A04": "招标文件公示",
    "A05": "澄清修改信息",
    "A06": "资格预审结果",
    "A07": "开标结果公示",
    "A08": "中标候选人公示",
    "A09": "中标结果公告",
    "A10": "合同信息公开",
    "A99": "其他",
    "1": "政府采购公告",
    "2": "中标公告",
    "3": "更正公告",
    "4": "竞争性磋商",
    "5": "单一来源",
}

# 排除关键词
EXCLUDE_KEYWORDS = ["监理", "设计", "咨询", "检测", "培训", "保洁", "绿化",
                     "广告", "升降平台", "液压升降", "吊篮", "施工升降机",
                     "升降作业", "货梯维修配件"]


def classify_business(title):
    """根据标题判断业务类型"""
    for biz_type, keywords in BUSINESS_RULES.items():
        for kw in keywords:
            if re.search(kw, title):
                return biz_type
    if "采购" in title:
        return "新梯采购"
    elif "维保" in title or "维护" in title:
        return "维保服务"
    elif "维修" in title or "改造" in title:
        return "维修改造"
    elif "加装" in title:
        return "旧梯加装"
    return "其他"


def extract_budget(title, content=""):
    """尝试从标题或内容中提取预算金额"""
    text = title + " " + content
    patterns = [
        r'预算[约金额]*[：:]*\s*(\d+\.?\d*)\s*万元',
        r'最高限价[：:]*\s*(\d+\.?\d*)\s*万元',
        r'控制价[：:]*\s*(\d+\.?\d*)\s*万元',
        r'招标金额[：:]*\s*(\d+\.?\d*)\s*万元',
        r'(\d+\.?\d*)\s*万元.*?电梯',
        r'采购预算[：:]*\s*(\d+\.?\d*)\s*万元',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            try:
                return float(match.group(1))
            except (ValueError, TypeError):
                pass
    return None


def is_excluded(title):
    """判断是否应排除"""
    for kw in EXCLUDE_KEYWORDS:
        if kw in title:
            return True
    return False


def search_api(keyword, start_date, end_date, page=0, page_size=50):
    """调用浙江省公共资源交易平台搜索API（单页）"""
    payload = {
        "token": "",
        "pn": page,
        "rn": page_size,
        "sdt": start_date,
        "edt": end_date,
        "wd": keyword,
        "inc_wd": "",
        "exc_wd": "",
        "fields": "title",
        "cnum": "001",
        "sort": '{"webdate":"0"}',
        "ssort": "title",
        "cl": 200,
        "terminal": "",
        "condition": [{
            "fieldName": "classcode",
            "fieldValue": "",
            "priority": 0,
            "operate": "eq"
        }]
    }

    try:
        r = requests.post(API_URL, json=payload, headers=HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()
        result = data.get("result", {})
        if isinstance(result, list):
            return {"records": result, "totalcount": len(result)}
        return result
    except Exception as e:
        print(f"    API请求失败(page={page}): {e}")
        return {}


def get_all_pages(keyword, start_date, end_date, max_pages=3):
    """分页获取所有结果"""
    all_records = []
    for page in range(max_pages):
        payload = {
            "token": "",
            "pn": page,
            "rn": 50,
            "sdt": start_date,
            "edt": end_date,
            "wd": keyword,
            "inc_wd": "",
            "exc_wd": "",
            "fields": "title",
            "cnum": "001",
            "sort": '{"webdate":"0"}',
            "ssort": "title",
            "cl": 200,
            "terminal": "",
            "condition": [{
                "fieldName": "classcode",
                "fieldValue": "",
                "priority": 0,
                "operate": "eq"
            }]
        }

        try:
            r = requests.post(API_URL, json=payload, headers=HEADERS, timeout=15)
            r.raise_for_status()
            data = r.json()
            result = data.get("result", {})
            if isinstance(result, list):
                records = result
            else:
                records = result.get("records", [])
                total = int(result.get("totalcount", 0))

            if not records:
                break

            all_records.extend(records)

            total = int(result.get("totalcount", 0))
            if len(all_records) >= total or len(records) < 50:
                break
        except Exception as e:
            print(f"    第{page+1}页请求失败: {e}")
            break

    return all_records


def build_detail_url(record):
    """构建详情页URL"""
    infourl = record.get("infourl", "")
    if infourl and infourl.startswith("http"):
        return infourl

    title = record.get("titlenew", "") or record.get("title", "")
    match = re.search(r'\[(A\d+)\]', title)
    if match:
        code = match.group(1)
        return f"https://ggzy.zj.gov.cn/jyxxgk/{code}.html"

    return "https://ggzy.zj.gov.cn/jyxxgk/list.html"


def generate_excel(all_bids, output_dir, filename_suffix=""):
    """生成带样式Excel报告"""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        df = pd.DataFrame(all_bids)

        # 按优先级排序
        priority = {"招标公告": 0, "资格预审公告": 1, "项目登记信息": 2, "招标文件公示": 3,
                   "中标候选人公示": 4, "中标结果公告": 5, "开标结果公示": 6, "澄清修改信息": 7,
                   "合同信息公开": 8, "其他": 9, "政府采购公告": 0, "中标公告": 5}
        df["_priority"] = df["公告类别"].map(priority).fillna(9)
        df = df.sort_values(["_priority", "发布时间"], ascending=[True, False])
        df = df.drop(columns=["_priority"])

        date_str = datetime.now().strftime("%Y-%m-%d")
        excel_path = output_dir / f"浙江电梯标讯_{date_str}{filename_suffix}.xlsx"
        df.to_excel(excel_path, index=False, engine="openpyxl")

        # 样式美化
        wb = load_workbook(excel_path)
        ws = wb.active

        header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
        header_font = Font(color="FFFFFF", size=11, bold=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        biz_colors = {
            "新梯采购": PatternFill(start_color="E6F1FB", end_color="E6F1FB", fill_type="solid"),
            "维保服务": PatternFill(start_color="EAF3DE", end_color="EAF3DE", fill_type="solid"),
            "维修改造": PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid"),
            "旧梯加装": PatternFill(start_color="FAECE7", end_color="FAECE7", fill_type="solid"),
        }

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            biz_type = row[4].value
            if biz_type in biz_colors:
                for cell in row:
                    if cell.column != 1:
                        cell.fill = biz_colors[biz_type]

        col_widths = {"A": 6, "B": 18, "C": 14, "D": 55, "E": 12, "F": 12, "G": 50, "H": 15}
        for col, width in col_widths.items():
            if col in ws.column_dimensions:
                ws.column_dimensions[col].width = width

        wb.save(excel_path)
        return excel_path

    except ImportError:
        date_str = datetime.now().strftime("%Y-%m-%d")
        csv_path = output_dir / f"浙江电梯标讯_{date_str}{filename_suffix}.csv"
        import csv
        if all_bids:
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=all_bids[0].keys())
                writer.writeheader()
                writer.writerows(all_bids)
        return csv_path


def collect(keywords, days, output_dir):
    """主采集逻辑"""
    today = datetime.now()
    date_str = today.strftime("%Y-%m-%d")
    start_date = (today - timedelta(days=days)).strftime("%Y-%m-%d")
    end_date = date_str

    print(f"\n{'='*60}")
    print(f"  浙江省电梯招投标信息自动采集")
    print(f"  采集日期: {date_str}")
    print(f"  采集范围: 最近 {days} 天 ({start_date} ~ {end_date})")
    print(f"  数据源: 浙江省公共资源交易平台 (ggzy.zj.gov.cn)")
    print(f"{'='*60}\n")

    output_dir.mkdir(exist_ok=True)
    all_bids = []
    seen_ids = set()

    for keyword in keywords:
        print(f"  搜索关键词: '{keyword}' ...", end=" ", flush=True)

        records = get_all_pages(keyword, start_date, end_date, max_pages=3)

        added = 0
        for rec in records:
            title_new = rec.get("titlenew", "") or rec.get("title", "")
            title_raw = rec.get("title", "")
            webdate = rec.get("webdate", "")
            infod = rec.get("infod", "")
            infoa = rec.get("infoa", "")
            content = rec.get("content", "")

            record_id = rec.get("id", "") or title_new[:50]
            if record_id in seen_ids:
                continue
            seen_ids.add(record_id)

            if is_excluded(title_new):
                continue

            biz_type = classify_business(title_new)
            category = CATEGORY_MAP.get(infoa, "其他")
            budget = extract_budget(title_new, content)
            detail_url = build_detail_url(rec)

            bid = {
                "序号": len(all_bids) + 1,
                "发布时间": webdate[:10] if len(webdate) >= 10 else webdate,
                "地区": infod,
                "项目名称": title_new,
                "业务类型": biz_type,
                "公告类别": category,
                "详情链接": detail_url,
                "预算(万元)": budget if budget else "",
            }
            all_bids.append(bid)
            added += 1

        print(f"新增 {added} 条有效标讯")

    all_bids.sort(key=lambda x: x.get("发布时间", ""), reverse=True)
    for i, bid in enumerate(all_bids, 1):
        bid["序号"] = i

    biz_counts = {}
    for bid in all_bids:
        bt = bid["业务类型"]
        biz_counts[bt] = biz_counts.get(bt, 0) + 1

    print(f"\n{'='*60}")
    print(f"  采集完成!")
    print(f"  总计: {len(all_bids)} 条有效标讯 (最近 {days} 天)")
    print(f"  分类统计:")
    for bt, count in biz_counts.items():
        print(f"    - {bt}: {count} 条")
    print(f"{'='*60}\n")

    if all_bids:
        excel_path = generate_excel(all_bids, output_dir)
        print(f"  Excel已保存: {excel_path}")
    else:
        print("  未采集到有效标讯")

    return all_bids


def main():
    parser = argparse.ArgumentParser(
        description="浙江省电梯招投标信息自动采集工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python collector.py                    # 采集最近7天，输出到 ./output
  python collector.py --days 3           # 采集最近3天
  python collector.py --output ./my_data # 输出到指定目录
  python collector.py --keywords "电梯,自动扶梯" # 自定义关键词
        """
    )
    parser.add_argument("--days", type=int, default=7,
                        help="采集最近几天的标讯 (默认: 7)")
    parser.add_argument("--output", type=str, default="./output",
                        help="输出目录 (默认: ./output)")
    parser.add_argument("--keywords", type=str, default=None,
                        help="搜索关键词，逗号分隔 (默认: 电梯,电梯采购,电梯维保,电梯维修,电梯改造,加装电梯)")
    args = parser.parse_args()

    keywords = [k.strip() for k in args.keywords.split(",")] if args.keywords else DEFAULT_KEYWORDS
    output_dir = Path(args.output)

    collect(keywords, args.days, output_dir)


if __name__ == "__main__":
    main()
